import os
import sys
import json
import email
import hashlib
import time
import csv
import certifi
import os
from pathlib import Path
from datetime import datetime, timedelta
from dotenv import load_dotenv
from email import policy
from email.parser import BytesParser
from azure.core.credentials import AzureKeyCredential
from azure.ai.documentintelligence import DocumentIntelligenceClient
from azure.ai.documentintelligence.models import AnalyzeDocumentRequest
from openai import AzureOpenAI
from azure.core.exceptions import HttpResponseError, ServiceRequestError
from io import BytesIO, StringIO
from urllib.parse import urlparse, unquote
os.environ['REQUESTS_CA_BUNDLE'] = certifi.where()
os.environ['SSL_CERT_FILE'] = certifi.where()

try:
    from azure.storage.blob import BlobServiceClient, ContainerClient, BlobClient
    BLOB_STORAGE_AVAILABLE = True
except ImportError:
    BLOB_STORAGE_AVAILABLE = False
    print("⚠️  WARNING: azure-storage-blob not installed. Blob storage support disabled.")
    print("   Install with: pip install azure-storage-blob")

# Check for openpyxl (REQUIRED for Excel support)
try:
    import openpyxl
    from openpyxl import load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️  WARNING: openpyxl not installed. Excel file support disabled.")
    print("   Install with: pip install openpyxl")

# Check for PDF/DOCX merging libraries
try:
    from PyPDF2 import PdfMerger, PdfReader
    from docx2pdf import convert as docx_to_pdf_convert
    import tempfile
    MERGE_AVAILABLE = True
except ImportError:
    MERGE_AVAILABLE = False
    print("⚠️  WARNING: PyPDF2 or docx2pdf not installed. Single API call optimization disabled.")
    print("   Install with: pip install PyPDF2 docx2pdf")

print()

# Global timing tracker
timing_data = {
    "process_start_time": None,
    "process_end_time": None,
    "folder_scan_start": None,
    "folder_scan_end": None,
    "file_detection_complete": None,
    "parsing_start_time": None,
    "parsing_complete_time": None,
    "doc_intelligence_start": None,
    "doc_intelligence_end": None,
    "ai_extraction_start": None,
    "ai_extraction_end": None,
    "json_save_time": None,
    "blob_download_start": None,
    "blob_download_end": None,
    "file_timings": {}
}

# Global API call counter
api_call_tracker = {
    "document_intelligence": 0,
    "azure_openai": 0
}

def print_timestamp(message):
    """Print message with timestamp."""
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
    print(f"⏱️  [{timestamp}] {message}")

def load_and_validate_env():
    """Load environment variables and validate they exist."""
    print("Loading environment variables...")
    
    env_path = Path(__file__).parent / ".env"
    load_dotenv(dotenv_path=env_path, override=True, verbose=True)
    
    required_vars = {
        "DOC_ENDPOINT": os.getenv("DOC_ENDPOINT"),
        "DOC_KEY": os.getenv("DOC_KEY"),
        "AZURE_OPENAI_ENDPOINT": os.getenv("AZURE_OPENAI_ENDPOINT"),
        "AZURE_OPENAI_KEY": os.getenv("AZURE_OPENAI_KEY")
    }
    
    # Optional blob storage variables
    blob_vars = {
        "AZURE_STORAGE_CONNECTION_STRING": os.getenv("AZURE_STORAGE_CONNECTION_STRING"),
        "AZURE_STORAGE_ACCOUNT_NAME": os.getenv("AZURE_STORAGE_ACCOUNT_NAME"),
        "AZURE_STORAGE_ACCOUNT_KEY": os.getenv("AZURE_STORAGE_ACCOUNT_KEY")
    }
    
    missing_vars = [var for var, value in required_vars.items() if not value]
    
    if missing_vars:
        print("\n❌ ERROR: Missing required environment variables:")
        for var in missing_vars:
            print(f"   - {var}")
        return None
    
    print("✅ All required environment variables loaded successfully")
    
    # Check blob storage config
    if all(blob_vars.values()) and BLOB_STORAGE_AVAILABLE:
        print("✅ Azure Blob Storage configured and available")
        required_vars.update(blob_vars)
    else:
        print("⚠️  Azure Blob Storage not fully configured (optional)")
    
    print()
    return required_vars

def sanitize_path(path_input):
    """Remove quotes and whitespace from path input."""
    return path_input.strip().strip('"').strip("'")

def calculate_md5(file_path_or_data):
    """Calculate MD5 hash of a file or data."""
    md5_hash = hashlib.md5()
    if isinstance(file_path_or_data, bytes):
        md5_hash.update(file_path_or_data)
    else:
        with open(file_path_or_data, 'rb') as f:
            for chunk in iter(lambda: f.read(4096), b""):
                md5_hash.update(chunk)
    return md5_hash.hexdigest()

def detect_file_type(file_path):
    """Detect file type from extension."""
    ext = Path(file_path).suffix.lower()
    if ext == '.eml':
        return 'EML'
    elif ext == '.pdf':
        return 'PDF'
    elif ext in ['.docx', '.doc']:
        return 'DOCX'
    elif ext in ['.xlsx', '.xls']:
        return 'EXCEL'
    elif ext == '.csv':
        return 'CSV'
    else:
        return None

def separate_pdf_and_other_files(file_paths):
    """Separate PDF files from other file types."""
    pdf_files = []
    other_files = []
    
    for file_path in file_paths:
        file_type = detect_file_type(file_path)
        if file_type == 'PDF':
            pdf_files.append(file_path)
        else:
            other_files.append(file_path)
    
    return pdf_files, other_files
    
def list_blob_containers(connection_string):
    """List all available blob containers."""
    try:
        blob_service_client = BlobServiceClient.from_connection_string(connection_string)
        containers = blob_service_client.list_containers()
        container_list = [container.name for container in containers]
        return container_list
    except Exception as e:
        print(f"❌ Error listing containers: {e}")
        return []

def list_blob_items_at_path(connection_string, container_name, prefix=""):
    """List folders and files at a specific blob path."""
    try:
        blob_service_client = BlobServiceClient.from_connection_string(connection_string)
        container_client = blob_service_client.get_container_client(container_name)
        
        # Get all blobs with the given prefix
        blob_list = container_client.walk_blobs(name_starts_with=prefix, delimiter='/')
        
        folders = []
        files = []
        
        for item in blob_list:
            if hasattr(item, 'name'):  # It's a blob (file)
                file_ext = Path(item.name).suffix.lower()
                if file_ext in {'.eml', '.pdf', '.docx', '.doc', '.xlsx', '.xls', '.csv'}:
                    files.append({
                        "name": item.name,
                        "display_name": Path(item.name).name,
                        "size": item.size,
                        "type": "file"
                    })
            else:  # It's a prefix (folder)
                folder_name = item.prefix if hasattr(item, 'prefix') else item.name
                folders.append({
                    "name": folder_name,
                    "display_name": folder_name.rstrip('/').split('/')[-1],
                    "type": "folder"
                })
        
        return folders, files
    
    except Exception as e:
        print(f"❌ Error listing blob items: {e}")
        return [], []

def interactive_blob_navigation(connection_string):
    """Interactive navigation through blob storage containers and folders."""
    print(f"\n{'='*100}")
    print("🌐 AZURE BLOB STORAGE NAVIGATION")
    print(f"{'='*100}\n")
    
    # Step 1: List and select container
    print("📦 Available Blob Containers:")
    containers = list_blob_containers(connection_string)
    
    if not containers:
        print("❌ No containers found")
        return None
    
    for idx, container in enumerate(containers, 1):
        print(f"   {idx}. {container}")
    
    while True:
        try:
            choice = input(f"\nSelect container (1-{len(containers)}) or 'q' to quit: ").strip()
            if choice.lower() == 'q':
                return None
            
            choice_idx = int(choice) - 1
            if 0 <= choice_idx < len(containers):
                selected_container = containers[choice_idx]
                break
            else:
                print(f"⚠️  Please enter a number between 1 and {len(containers)}")
        except ValueError:
            print("⚠️  Please enter a valid number")
    
    print(f"\n✅ Selected container: {selected_container}")
    
    # Step 2: Navigate through folders
    current_prefix = ""
    selected_files = []
    
    while True:
        print(f"\n{'='*100}")
        print(f"📂 Current path: {selected_container}/{current_prefix if current_prefix else '(root)'}")
        print(f"{'='*100}\n")
        
        folders, files = list_blob_items_at_path(connection_string, selected_container, current_prefix)
        
        if not folders and not files:
            print("❌ No items found at this path")
            if current_prefix:
                current_prefix = '/'.join(current_prefix.rstrip('/').split('/')[:-1])
                if current_prefix:
                    current_prefix += '/'
                continue
            else:
                return None
        
        # Display folders
        if folders:
            print("📁 Folders:")
            for idx, folder in enumerate(folders, 1):
                print(f"   {idx}. 📁 {folder['display_name']}/")
        
        # Display files
        if files:
            print("\n📄 Files:")
            file_start_idx = len(folders) + 1
            for idx, file in enumerate(files, file_start_idx):
                size_mb = file['size'] / (1024 * 1024)
                print(f"   {idx}. 📄 {file['display_name']} ({size_mb:.2f} MB)")
        
        # Navigation options
        print(f"\n{'='*100}")
        print("Options:")
        print("  • Enter number(s) to select items (e.g., '1' or '1,3,5' or '1-5')")
        print("  • Enter 'all' to select all files in current folder")
        print("  • Enter 'back' to go up one level")
        print("  • Enter 'done' to finish selection")
        print("  • Enter 'q' to quit")
        print(f"{'='*100}")
        
        choice = input("\nYour choice: ").strip().lower()
        
        if choice == 'q':
            return None
        
        if choice == 'back':
            if current_prefix:
                # Go up one level
                current_prefix = '/'.join(current_prefix.rstrip('/').split('/')[:-1])
                if current_prefix:
                    current_prefix += '/'
            continue
        
        if choice == 'done':
            if selected_files:
                return {
                    "container": selected_container,
                    "files": selected_files
                }
            else:
                print("⚠️  No files selected yet")
                continue
        
        if choice == 'all':
            # Select all files in current folder
            for file in files:
                if file not in selected_files:
                    selected_files.append(file)
            print(f"✅ Added {len(files)} file(s) to selection")
            print(f"   Total selected: {len(selected_files)} file(s)")
            continue
        
        # Parse selection
        try:
            selected_indices = []
            
            # Handle ranges and comma-separated values
            parts = choice.split(',')
            for part in parts:
                part = part.strip()
                if '-' in part:
                    start, end = map(int, part.split('-'))
                    selected_indices.extend(range(start, end + 1))
                else:
                    selected_indices.append(int(part))
            
            # Process selections
            for idx in selected_indices:
                if 1 <= idx <= len(folders):
                    # Navigate into folder
                    folder = folders[idx - 1]
                    current_prefix = folder['name']
                    break  # Navigate into first selected folder
                elif len(folders) < idx <= len(folders) + len(files):
                    # Select file
                    file_idx = idx - len(folders) - 1
                    file = files[file_idx]
                    if file not in selected_files:
                        selected_files.append(file)
                        print(f"✅ Added: {file['display_name']}")
                else:
                    print(f"⚠️  Invalid selection: {idx}")
            
            if selected_files:
                print(f"\n📊 Total selected: {len(selected_files)} file(s)")
        
        except ValueError:
            print("⚠️  Invalid input format")
            continue

def is_blob_storage_url(path):
    """Check if path is an Azure Blob Storage URL."""
    path = path.lower()
    return (path.startswith("https://") and ".blob.core.windows.net" in path) or path.startswith("blob://")

def parse_blob_path(blob_path):
    """Parse blob storage path in multiple formats."""
    blob_path = blob_path.strip()
    
    # Handle full HTTPS URL
    if blob_path.startswith("https://"):
        try:
            parsed_url = urlparse(blob_path)
            decoded_path = unquote(parsed_url.path)
            path_parts = decoded_path.lstrip('/').split('/', 1)
            container_name = path_parts[0]
            prefix_path = path_parts[1] if len(path_parts) > 1 else ""
            
            print(f"   🔍 Parsed URL:")
            print(f"      Account: {parsed_url.netloc.split('.')[0]}")
            print(f"      Container: {container_name}")
            print(f"      Path: {prefix_path if prefix_path else '(root)'}")
            
            return container_name, prefix_path
        except Exception as e:
            print(f"⚠️  Error parsing blob URL: {e}")
            return None, None
    
    # Handle blob:// prefix
    elif blob_path.startswith("blob://"):
        blob_path = blob_path[7:]
    
    # Simple format: container/path
    parts = blob_path.split('/', 1)
    container_name = parts[0]
    prefix_path = parts[1] if len(parts) > 1 else ""
    
    return container_name, prefix_path

def scan_blob_storage_for_files(connection_string, container_name, prefix=""):
    """Scan Azure Blob Storage container for supported files."""
    if not BLOB_STORAGE_AVAILABLE:
        print("❌ ERROR: azure-storage-blob library not installed")
        return None
    
    print_timestamp(f"Scanning blob storage: {container_name}/{prefix if prefix else '(root)'}")
    timing_data["blob_download_start"] = time.time()
    
    supported_extensions = {'.eml', '.pdf', '.docx', '.doc', '.xlsx', '.xls', '.csv'}
    found_blobs = []
    
    try:
        blob_service_client = BlobServiceClient.from_connection_string(connection_string)
        container_client = blob_service_client.get_container_client(container_name)
        
        # List all blobs with the given prefix
        blob_list = container_client.list_blobs(name_starts_with=prefix)
        
        blob_count = 0
        for blob in blob_list:
            blob_name = blob.name
            blob_ext = Path(blob_name).suffix.lower()
            
            if blob_ext in supported_extensions:
                found_blobs.append({
                    "name": blob_name,
                    "size": blob.size,
                    "container": container_name,
                    "full_path": f"blob://{container_name}/{blob_name}"
                })
            blob_count += 1
        
        print(f"   ✅ Scanned {blob_count} blobs")
        print(f"   ✅ Found {len(found_blobs)} supported file(s)")
        
        if found_blobs:
            print(f"\n   📄 Files to download:")
            for idx, blob in enumerate(found_blobs[:10], 1):
                size_mb = blob['size'] / (1024 * 1024)
                print(f"      {idx}. {Path(blob['name']).name} ({size_mb:.2f} MB)")
            if len(found_blobs) > 10:
                print(f"      ... and {len(found_blobs) - 10} more files")
        print()
        
        return found_blobs
    
    except Exception as e:
        print(f"❌ ERROR scanning blob storage: {type(e).__name__}: {e}")
        return None

def download_blob_to_memory(connection_string, container_name, blob_name):
    """Download a blob file to memory."""
    try:
        blob_service_client = BlobServiceClient.from_connection_string(connection_string)
        blob_client = blob_service_client.get_blob_client(container=container_name, blob=blob_name)
        
        blob_data = blob_client.download_blob().readall()
        return blob_data
    
    except Exception as e:
        print(f"❌ Error downloading blob {blob_name}: {e}")
        return None

def download_blobs_to_temp_files(connection_string, blob_info_list):
    """Download blob files and return temporary file paths."""
    print_timestamp(f"Downloading {len(blob_info_list)} files from blob storage...")
    
    temp_files = []
    
    for idx, blob_info in enumerate(blob_info_list, 1):
        container = blob_info['container']
        blob_name = blob_info['name']
        size_mb = blob_info['size'] / (1024 * 1024)
        
        print(f"   [{idx}/{len(blob_info_list)}] Downloading: {Path(blob_name).name} ({size_mb:.2f} MB)")
        
        blob_data = download_blob_to_memory(connection_string, container, blob_name)
        
        if blob_data:
            # Create temporary file
            temp_dir = tempfile.gettempdir()
            temp_filename = Path(blob_name).name
            temp_path = Path(temp_dir) / f"blob_{int(time.time())}_{idx}_{temp_filename}"
            
            with open(temp_path, 'wb') as f:
                f.write(blob_data)
            
            temp_files.append({
                "temp_path": str(temp_path),
                "original_name": blob_name,
                "blob_info": blob_info
            })
            print(f"       ✅ Downloaded successfully")
        else:
            print(f"       ❌ Download failed")
    
    timing_data["blob_download_end"] = time.time()
    if timing_data["blob_download_start"]:
        duration = timing_data["blob_download_end"] - timing_data["blob_download_start"]
        total_size_mb = sum(b['size'] for b in blob_info_list) / (1024 * 1024)
        print(f"\n✅ Blob download completed in {duration:.3f} seconds ({total_size_mb:.2f} MB total)\n")
    
    return temp_files

# ==================== EXISTING FUNCTIONS (UNCHANGED) ====================

def scan_folder_for_files(folder_path):
    """Scan folder and subfolders for supported files."""
    print_timestamp(f"Scanning local folder: {folder_path}")
    timing_data["folder_scan_start"] = time.time()
    
    supported_extensions = {'.eml', '.pdf', '.docx', '.doc', '.xlsx', '.xls', '.csv'}
    found_files = []
    
    try:
        folder = Path(folder_path)
        
        if not folder.exists():
            print(f"❌ ERROR: Folder does not exist: {folder_path}")
            return None
        
        if not folder.is_dir():
            print(f"❌ ERROR: Path is not a folder: {folder_path}")
            return None
        
        try:
            list(folder.iterdir())
        except PermissionError:
            print(f"❌ ERROR: Access denied to folder: {folder_path}")
            return None
        
        for file_path in folder.rglob('*'):
            if file_path.is_file() and file_path.suffix.lower() in supported_extensions:
                try:
                    file_path.stat()
                    found_files.append(str(file_path))
                except PermissionError:
                    print(f"   ⚠️  Skipping (access denied): {file_path.name}")
                except Exception as e:
                    print(f"   ⚠️  Skipping (error): {file_path.name} - {e}")
        
        timing_data["folder_scan_end"] = time.time()
        scan_duration = timing_data["folder_scan_end"] - timing_data["folder_scan_start"]
        
        print(f"✅ Folder scan completed in {scan_duration:.3f} seconds")
        print(f"✅ Found {len(found_files)} supported file(s)\n")
        
        return found_files
    
    except Exception as e:
        print(f"❌ ERROR scanning folder: {type(e).__name__}: {e}")
        return None

def display_detected_files(file_paths):
    """Display all detected files in a formatted table."""
    print(f"\n{'='*100}")
    print("🔍 DETECTED FILES")
    print(f"{'='*100}")
    print(f"{'#':<5} {'Filename':<40} {'Type':<10} {'Size':<15} {'Path':<30}")
    print(f"{'-'*100}")
    
    for idx, file_path in enumerate(file_paths, 1):
        path = Path(file_path)
        file_type = detect_file_type(file_path)
        try:
            size = path.stat().st_size
            size_str = f"{size:,} bytes"
        except:
            size_str = "Unknown"
        
        filename = path.name[:38] + '..' if len(path.name) > 40 else path.name
        rel_path = str(path.parent)[:28] + '..' if len(str(path.parent)) > 30 else str(path.parent)
        
        print(f"{idx:<5} {filename:<40} {file_type:<10} {size_str:<15} {rel_path:<30}")
    
    print(f"{'='*100}\n")
    timing_data["file_detection_complete"] = time.time()

def parse_csv_file(csv_path_or_data, filename=None):
    """Parse CSV file and extract all data."""
    file_start = time.time()
    print_timestamp(f"Parsing CSV: {filename or csv_path_or_data}")
    
    try:
        if isinstance(csv_path_or_data, bytes):
            csv_text = csv_path_or_data.decode('utf-8', errors='ignore')
        else:
            with open(csv_path_or_data, 'r', encoding='utf-8', errors='ignore') as f:
                csv_text = f.read()
        
        csv_reader = csv.reader(StringIO(csv_text))
        rows = [row for row in csv_reader if any(cell.strip() for cell in row)]
        
        headers = rows[0] if rows else []
        data_rows = rows[1:] if len(rows) > 1 else []
        
        csv_text_repr = f"---{filename or 'CSV_FILE'}--- ---CSV---\n\n"
        csv_text_repr += "HEADERS: " + " | ".join(headers) + "\n\n"
        
        for idx, row in enumerate(data_rows[:100], start=1):
            csv_text_repr += f"Row {idx}: " + " | ".join(row) + "\n"
        
        if len(data_rows) > 100:
            csv_text_repr += f"\n... and {len(data_rows) - 100} more rows\n"
        
        csv_data = {
            "headers": headers,
            "total_rows": len(data_rows),
            "total_columns": len(headers),
            "data_rows": data_rows,
            "text": csv_text_repr,
            "raw_csv": csv_text[:10000]
        }
        
        file_end = time.time()
        timing_data["file_timings"][filename or "CSV"] = {
            "start": file_start,
            "end": file_end,
            "duration": file_end - file_start
        }
        
        print(f"   ✅ Parsed: {len(headers)} columns, {len(data_rows)} rows ({file_end - file_start:.3f}s)\n")
        return csv_data
    
    except Exception as e:
        print(f"   ❌ Error: {type(e).__name__}: {e}\n")
        return None

def parse_excel_file(excel_path_or_data, filename=None):
    """Parse Excel file and extract all sheet data."""
    if not EXCEL_AVAILABLE:
        print(f"❌ ERROR: Cannot parse Excel - openpyxl not installed")
        return None
    
    file_start = time.time()
    print_timestamp(f"Parsing Excel: {filename or excel_path_or_data}")
    
    try:
        if isinstance(excel_path_or_data, bytes):
            wb = load_workbook(BytesIO(excel_path_or_data), data_only=True)
        else:
            wb = load_workbook(excel_path_or_data, data_only=True)
        
        excel_data = {
            "sheets": [],
            "sheet_names": wb.sheetnames,
            "total_sheets": len(wb.sheetnames)
        }
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            rows_data = []
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                row_list = [str(cell) if cell is not None else "" for cell in row]
                if any(row_list):
                    rows_data.append({"row_number": row_idx, "cells": row_list})
            
            sheet_text = f"--- SHEET: {sheet_name} ({len(rows_data)} rows) ---\n"
            for row in rows_data[:100]:
                sheet_text += " | ".join(row["cells"]) + "\n"
            
            excel_data["sheets"].append({
                "sheet_name": sheet_name,
                "rows": rows_data,
                "text": sheet_text,
                "row_count": len(rows_data),
                "column_count": sheet.max_column
            })
        
        file_end = time.time()
        timing_data["file_timings"][filename or "EXCEL"] = {
            "start": file_start,
            "end": file_end,
            "duration": file_end - file_start
        }
        
        print(f"   ✅ Parsed: {len(excel_data['sheets'])} sheets ({file_end - file_start:.3f}s)\n")
        return excel_data
    
    except Exception as e:
        print(f"   ❌ Error: {type(e).__name__}: {e}\n")
        return None

def parse_eml_file(eml_path):
    """Parse EML file and extract email metadata and attachments."""
    file_start = time.time()
    print_timestamp(f"Parsing EML: {Path(eml_path).name}")
    
    try:
        with open(eml_path, 'rb') as f:
            msg = BytesParser(policy=policy.default).parse(f)
        
        eml_data = {
            "filename": Path(eml_path).name,
            "from": msg.get('From', ''),
            "to": msg.get('To', ''),
            "cc": msg.get('Cc', ''),
            "subject": msg.get('Subject', ''),
            "date": msg.get('Date', ''),
            "message_id": msg.get('Message-ID', ''),
            "body_text": "",
            "body_html": "",
            "attachments": []
        }
        
        if msg.is_multipart():
            for part in msg.walk():
                content_type = part.get_content_type()
                content_disposition = str(part.get('Content-Disposition', ''))
                
                if 'attachment' not in content_disposition:
                    if content_type == 'text/plain':
                        try:
                            eml_data["body_text"] += part.get_content()
                        except:
                            pass
                    elif content_type == 'text/html':
                        try:
                            eml_data["body_html"] += part.get_content()
                        except:
                            pass
                else:
                    filename = part.get_filename()
                    if filename:
                        file_ext = Path(filename).suffix.lower()
                        # Exclude PDFs from EML attachments - they'll be handled separately
                        if file_ext in ['.docx', '.doc', '.xlsx', '.xls', '.csv']:
                            file_data = part.get_payload(decode=True)
                            detected_type = detect_file_type(filename)
                            
                            if detected_type == 'EXCEL' and not EXCEL_AVAILABLE:
                                print(f"   ⚠️  Skipping {filename} - openpyxl not installed")
                                continue
                            
                            eml_data["attachments"].append({
                                "filename": filename,
                                "content_type": content_type,
                                "data": file_data,
                                "size": len(file_data),
                                "md5sum": calculate_md5(file_data),
                                "file_type": detected_type,
                                "source_eml": eml_data["filename"]
                            })
        else:
            try:
                eml_data["body_text"] = msg.get_content()
            except:
                pass
        
        file_end = time.time()
        timing_data["file_timings"][Path(eml_path).name] = {
            "start": file_start,
            "end": file_end,
            "duration": file_end - file_start
        }
        
        print(f"   ✅ Parsed: {len(eml_data['attachments'])} attachments ({file_end - file_start:.3f}s)\n")
        return eml_data
    
    except Exception as e:
        print(f"   ❌ Error: {type(e).__name__}: {e}\n")
        return None

def collect_all_documents(file_paths):
    """Collect all documents from file paths including EML attachments. Excludes PDFs."""
    print_timestamp("Starting document collection (excluding PDFs)")
    timing_data["parsing_start_time"] = time.time()
    
    all_eml_data = []
    all_documents = []
    
    for file_path in file_paths:
        file_type = detect_file_type(file_path)
        if not file_type or file_type == 'PDF':  # Skip PDFs
            continue
        
        if file_type == 'EML':
            eml_data = parse_eml_file(file_path)
            if eml_data:
                all_eml_data.append(eml_data)
                
                for att in eml_data['attachments']:
                    all_documents.append({
                        "filename": att['filename'],
                        "file_type": att['file_type'],
                        "data": att['data'],
                        "size": att['size'],
                        "md5sum": att['md5sum'],
                        "source": f"Attachment from {eml_data['filename']}"
                    })
        else:
            try:
                with open(file_path, 'rb') as f:
                    file_data = f.read()
                
                all_documents.append({
                    "filename": Path(file_path).name,
                    "file_type": file_type,
                    "data": file_data,
                    "size": len(file_data),
                    "md5sum": calculate_md5(file_data),
                    "source": "Direct file"
                })
            except Exception as e:
                print(f"   ❌ Error loading {Path(file_path).name}: {e}")
    
    timing_data["parsing_complete_time"] = time.time()
    parsing_duration = timing_data["parsing_complete_time"] - timing_data["parsing_start_time"]
    
    print(f"\n{'='*100}")
    print(f"✅ Document collection completed in {parsing_duration:.3f} seconds")
    print(f"   EML Files: {len(all_eml_data)}")
    print(f"   Total Documents (non-PDF): {len(all_documents)}")
    print(f"{'='*100}\n")
    
    return all_eml_data, all_documents

def analyze_documents_single_call(all_eml_data, all_documents, doc_endpoint, doc_key):
    """Analyze documents with robust error handling and fallback options."""
    print_timestamp("Starting Document Intelligence analysis")
    timing_data["doc_intelligence_start"] = time.time()
    
    try:
        # Configure SSL and retry settings
        import ssl
        import urllib3
        from urllib3.util.retry import Retry
        from requests.adapters import HTTPAdapter
        from requests import Session
        from azure.core.pipeline.transport import RequestsTransport
        from azure.core.pipeline.policies import RetryPolicy
        
        # Disable SSL warnings for development
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        
        # Create session with retry logic and SSL configuration
        session = Session()
        session.verify = certifi.where()  # Use certifi certificates
        
        # Configure retry strategy
        retry_strategy = Retry(
            total=3,
            backoff_factor=2,
            status_forcelist=[429, 500, 502, 503, 504],
            allowed_methods=["POST"]
        )
        
        adapter = HTTPAdapter(max_retries=retry_strategy)
        session.mount("https://", adapter)
        session.mount("http://", adapter)
        
        # Increase timeout
        transport = RequestsTransport(session=session)
        
        # Create client with custom transport
        client = DocumentIntelligenceClient(
            endpoint=doc_endpoint,
            credential=AzureKeyCredential(doc_key),
            transport=transport
        )
        
        extracted_data = {"EML_FILES": [], "DOCUMENTS": []}
        
        # Process EML files (text-based, no API call needed)
        print_timestamp(f"Processing {len(all_eml_data)} EML files...")
        for eml_data in all_eml_data:
            eml_text = f"---{eml_data['filename']}--- ---EML---\n\n"
            eml_text += f"From: {eml_data['from']}\nTo: {eml_data['to']}\n"
            eml_text += f"Subject: {eml_data['subject']}\nDate: {eml_data['date']}\n\n"
            eml_text += f"{eml_data['body_text'] or eml_data['body_html'][:5000]}\n"
            
            extracted_data["EML_FILES"].append({
                "filename": eml_data['filename'],
                "text": eml_text,
                "metadata": {
                    "from": eml_data['from'],
                    "to": eml_data['to'],
                    "subject": eml_data['subject'],
                    "date": eml_data['date']
                }
            })
        
        # Categorize documents (no PDFs here)
        docx_docs = [d for d in all_documents if d['file_type'] == 'DOCX']
        excel_docs = [d for d in all_documents if d['file_type'] == 'EXCEL']
        csv_docs = [d for d in all_documents if d['file_type'] == 'CSV']
        
        print_timestamp(f"Found {len(docx_docs)} DOCX, {len(excel_docs)} Excel, {len(csv_docs)} CSV documents")
        
        # Process DOCX documents
        if docx_docs:
            print_timestamp(f"Processing {len(docx_docs)} DOCX documents individually...")
            
            for idx, doc in enumerate(docx_docs, 1):
                try:
                    print(f"   [{idx}/{len(docx_docs)}] Processing {doc['filename']}...")
                    
                    # Convert DOCX to PDF if needed
                    doc_data = doc['data']
                    if doc['file_type'] == 'DOCX':
                        try:
                            with tempfile.NamedTemporaryFile(suffix='.docx', delete=False) as docx_temp:
                                docx_temp.write(doc['data'])
                                docx_temp_path = docx_temp.name
                            
                            pdf_temp_path = docx_temp_path.replace('.docx', '.pdf')
                            docx_to_pdf_convert(docx_temp_path, pdf_temp_path)
                            
                            with open(pdf_temp_path, 'rb') as f:
                                doc_data = f.read()
                            
                            # Cleanup
                            os.unlink(docx_temp_path)
                            os.unlink(pdf_temp_path)
                            
                        except Exception as conv_err:
                            print(f"      ⚠️  DOCX conversion failed: {conv_err}")
                            print(f"      ℹ️  Using raw DOCX data (may have limited extraction)")
                    
                    # Analyze individual document
                    api_call_tracker["document_intelligence"] += 1
                    
                    poller = client.begin_analyze_document(
                        "prebuilt-layout",
                        AnalyzeDocumentRequest(bytes_source=doc_data)
                    )
                    
                    result = poller.result()
                    
                    # Extract text
                    doc_text = f"---{doc['filename']}--- ---{doc['file_type']}---\n\n"
                    pages_text = []
                    
                    if result.pages:
                        for page_idx, page in enumerate(result.pages, start=1):
                            page_content = ""
                            if page.lines:
                                for line in page.lines:
                                    page_content += line.content + "\n"
                            pages_text.append({"page_number": page_idx, "content": page_content})
                            doc_text += page_content + "\n"
                    
                    extracted_data["DOCUMENTS"].append({
                        "filename": doc['filename'],
                        "file_type": doc['file_type'],
                        "source": doc.get('source', 'Unknown'),
                        "text": doc_text,
                        "pages_text": pages_text,
                        "page_range": f"Pages 1-{len(pages_text)}"
                    })
                    
                    print(f"      ✅ Processed successfully ({len(pages_text)} pages)")
                    
                except Exception as e:
                    print(f"      ❌ Error: {type(e).__name__}: {e}")
                    
                    extracted_data["DOCUMENTS"].append({
                        "filename": doc['filename'],
                        "file_type": doc['file_type'],
                        "source": doc.get('source', 'Unknown'),
                        "text": f"---{doc['filename']}--- ---{doc['file_type']}---\n\nERROR: Failed to process this document.\nError: {str(e)[:200]}",
                        "pages_text": [],
                        "page_range": "N/A",
                        "extraction_error": str(e)
                    })
        
        # Process Excel documents (no API call needed)
        print_timestamp(f"Processing {len(excel_docs)} Excel documents...")
        for doc in excel_docs:
            excel_data = parse_excel_file(doc['data'], doc['filename'])
            if excel_data:
                combined_text = f"---{doc['filename']}--- ---EXCEL---\n\n"
                for sheet in excel_data['sheets']:
                    combined_text += sheet['text'] + "\n\n"
                
                extracted_data["DOCUMENTS"].append({
                    "filename": doc['filename'],
                    "file_type": 'EXCEL',
                    "source": doc.get('source', 'Unknown'),
                    "text": combined_text,
                    "excel_data": excel_data
                })
        
        # Process CSV documents (no API call needed)
        print_timestamp(f"Processing {len(csv_docs)} CSV documents...")
        for doc in csv_docs:
            csv_data = parse_csv_file(doc['data'], doc['filename'])
            if csv_data:
                extracted_data["DOCUMENTS"].append({
                    "filename": doc['filename'],
                    "file_type": 'CSV',
                    "source": doc.get('source', 'Unknown'),
                    "text": csv_data['text'],
                    "csv_data": csv_data
                })
        
        timing_data["doc_intelligence_end"] = time.time()
        duration = timing_data["doc_intelligence_end"] - timing_data["doc_intelligence_start"]
        
        print(f"\n{'='*100}")
        print(f"✅ Document Intelligence completed ({duration:.3f}s, {api_call_tracker['document_intelligence']} API calls)")
        print(f"   📧 EML Files: {len(extracted_data['EML_FILES'])}")
        print(f"   📄 Documents: {len(extracted_data['DOCUMENTS'])}")
        print(f"{'='*100}\n")
        
        return extracted_data
    
    except Exception as e:
        print(f"❌ Critical Error in analyze_documents_single_call: {type(e).__name__}: {e}\n")
        import traceback
        traceback.print_exc()
        
        # Return partial data if available
        if 'extracted_data' in locals():
            return extracted_data
        
        return None

def merge_all_pdfs_and_docx(documents):
    """Merge ALL PDF and DOCX files into single PDF."""
    if not MERGE_AVAILABLE:
        return None
    
    print_timestamp("Starting document merging")
    
    try:
        merger = PdfMerger()
        temp_files = []
        doc_map = []
        current_page = 1
        
        for doc in documents:
            if doc['file_type'] not in ['PDF', 'DOCX']:
                continue
            
            try:
                if doc['file_type'] == 'DOCX':
                    with tempfile.NamedTemporaryFile(suffix='.docx', delete=False) as docx_temp:
                        docx_temp.write(doc['data'])
                        docx_temp_path = docx_temp.name
                    temp_files.append(docx_temp_path)
                    
                    pdf_temp_path = docx_temp_path.replace('.docx', '.pdf')
                    docx_to_pdf_convert(docx_temp_path, pdf_temp_path)
                    temp_files.append(pdf_temp_path)
                    
                    with open(pdf_temp_path, 'rb') as f:
                        pdf_data = f.read()
                    
                    pdf_reader = PdfReader(BytesIO(pdf_data))
                    page_count = len(pdf_reader.pages)
                    merger.append(BytesIO(pdf_data))
                    
                elif doc['file_type'] == 'PDF':
                    pdf_reader = PdfReader(BytesIO(doc['data']))
                    page_count = len(pdf_reader.pages)
                    merger.append(BytesIO(doc['data']))
                
                doc_map.append({
                    "filename": doc['filename'],
                    "file_type": doc['file_type'],
                    "source": doc.get('source', 'Unknown'),
                    "start_page": current_page,
                    "end_page": current_page + page_count - 1,
                    "page_count": page_count
                })
                current_page += page_count
                
            except Exception as e:
                print(f"   ⚠️  Could not merge {doc['filename']}: {e}")
                continue
        
        if not doc_map:
            return None
        
        merged_output = BytesIO()
        merger.write(merged_output)
        merged_output.seek(0)
        merged_pdf_data = merged_output.read()
        merger.close()
        
        for temp_file in temp_files:
            try:
                os.unlink(temp_file)
            except:
                pass
        
        print(f"   ✅ Merged {len(doc_map)} documents ({current_page - 1} pages)\n")
        
        return {
            "merged_pdf_data": merged_pdf_data,
            "document_map": doc_map,
            "total_pages": current_page - 1
        }
    
    except Exception as e:
        print(f"   ❌ Merge error: {e}\n")
        return None

def extract_fields_with_file_sections(extracted_data, openai_endpoint, openai_key):
    """Extract fields with proper token limit management."""
    print_timestamp("Starting Token-Optimized Batch Extraction")
    timing_data["ai_extraction_start"] = time.time()
    
    try:
        client = AzureOpenAI(
            azure_endpoint=openai_endpoint,
            api_key=openai_key,
            api_version=os.getenv("AZURE_OPENAI_API_VERSION", "2024-02-15-preview")
        )
        
        deployment_name = os.getenv("AZURE_OPENAI_DEPLOYMENT", "gpt-4o")
        
        # Track file information
        processed_files_info = {}
        document_content_map = {}
        
        print(f"   📋 Processing {len(extracted_data.get('EML_FILES', []))} EML files...")
        print(f"   📋 Processing {len(extracted_data.get('DOCUMENTS', []))} documents...")
        
        # Process EML files
        for eml in extracted_data.get('EML_FILES', []):
            eml_filename = eml['filename']
            
            processed_files_info[eml_filename] = {
                "file_type": "EML",
                "is_attachment": False,
                "has_attachments": False,
                "attachment_count": 0
            }
            
            document_content_map[eml_filename] = {
                "type": "EML",
                "content": eml['text'][:10000],
                "attachments": []
            }
        
        # Process documents
        for doc in extracted_data.get('DOCUMENTS', []):
            doc_filename = doc['filename']
            doc_type = doc['file_type']
            source = doc.get('source', 'Unknown')
            doc_content = doc.get('text', '')[:10000]
            
            # Check if attachment
            if 'Attachment from' in source:
                parent_eml = source.replace('Attachment from ', '').strip()
                
                if parent_eml in processed_files_info:
                    processed_files_info[parent_eml]["has_attachments"] = True
                    processed_files_info[parent_eml]["attachment_count"] += 1
                    
                    processed_files_info[doc_filename] = {
                        "file_type": doc_type,
                        "is_attachment": True,
                        "parent_file": parent_eml
                    }
                    
                    att_num = processed_files_info[parent_eml]["attachment_count"]
                    
                    if parent_eml in document_content_map:
                        document_content_map[parent_eml]["attachments"].append({
                            "filename": doc_filename,
                            "type": doc_type,
                            "content": doc_content,
                            "att_num": att_num,
                            "page_range": doc.get('page_range', 'N/A')
                        })
            else:
                # Standalone document
                processed_files_info[doc_filename] = {
                    "file_type": doc_type,
                    "is_attachment": False
                }
                
                document_content_map[doc_filename] = {
                    "type": doc_type,
                    "content": doc_content,
                    "page_range": doc.get('page_range', 'N/A'),
                    "attachments": []
                }
        
        # OPTIMIZED BATCH PROCESSING
        document_list = list(document_content_map.keys())
        batch_size = 2
        all_extracted_data = {}
        
        print(f"   📦 Processing {len(document_list)} documents in batches of {batch_size}")
        print(f"   ⚙️  Max tokens per batch: 15000 (model limit: 16384)\n")
        
        for batch_start in range(0, len(document_list), batch_size):
            batch_end = min(batch_start + batch_size, len(document_list))
            batch_docs = document_list[batch_start:batch_end]
            batch_num = batch_start // batch_size + 1
            total_batches = (len(document_list) - 1) // batch_size + 1
            
            print(f"   📄 Batch {batch_num}/{total_batches}: Processing {len(batch_docs)} document(s)")
            
            # Build COMPACT prompt
            combined_content = ""
            
            for doc_idx, main_filename in enumerate(batch_docs, 1):
                doc_info = document_content_map[main_filename]
                
                combined_content += f"\n{'='*80}\n"
                combined_content += f"DOCUMENT {doc_idx}: {main_filename} [{doc_info['type']}]\n"
                combined_content += f"{'='*80}\n"
                
                if doc_info['type'] == 'EML':
                    combined_content += doc_info['content'] + "\n"
                    
                    if doc_info['attachments']:
                        for att in doc_info['attachments']:
                            combined_content += f"\n--- ATTACHMENT: {att['filename']} [{att['type']}] ---\n"
                            combined_content += att['content'] + "\n"
                else:
                    combined_content += doc_info['content'] + "\n"
            
            # Strict content size limit
            if len(combined_content) > 50000:
                print(f"      ⚠️  Content truncated: {len(combined_content):,} → 50,000 chars")
                combined_content = combined_content[:50000]
            
            print(f"      📊 Content size: {len(combined_content):,} characters")
            
            # COMPACT SYSTEM PROMPT
            system_prompt = """Extract ALL fields from documents. Return JSON only.

Rules:
1. Extract 20+ fields per document
2. Use format: {"value": "...", "confidence": 0.0-1.0, "page": "..."}
3. Field names: Use_Underscores (Email_From, Invoice_Number, etc.)
4. Include null for missing fields
5. Return ALL documents provided

No markdown. No explanations. JSON only."""

            # COMPACT USER PROMPT
            files_list = ", ".join([f"{i+1}. {fname}" for i, fname in enumerate(batch_docs)])
            
            user_prompt = f"""Extract fields from {len(batch_docs)} document(s): {files_list}

REQUIRED: Return ALL {len(batch_docs)} documents in JSON.

For emails: Email_From, Email_To, Email_CC, Email_Subject, Email_Date, etc.
For PDFs: Document_Title, Document_Date, Company_Name, Amount, Address, Phone, etc.
For Excel: Sheet_Name, Column_Headers, Total_Records, data fields, etc.

JSON Format:
{{
  "filename.eml": {{
    "EML_FILE_DATA": {{
      "Field_Name": {{"value": "...", "confidence": 0.95, "page": "Email Header"}},
      ...minimum 20 fields...
    }},
    "ATTACHMENT_1_PDF_FILE_DATA": {{...20+ fields...}}
  }},
  "document.pdf": {{
    "PDF_FILE_DATA": {{...20+ fields...}}
  }}
}}

Documents:
{combined_content}

Return JSON for ALL {len(batch_docs)} files."""

            # API CALL
            print(f"      ⚡ Calling Azure OpenAI API...")
            api_call_tracker["azure_openai"] += 1
            
            try:
                response = client.chat.completions.create(
                    model=deployment_name,
                    messages=[
                        {"role": "system", "content": system_prompt},
                        {"role": "user", "content": user_prompt}
                    ],
                    temperature=0.05,
                    max_tokens=15000,
                    response_format={"type": "json_object"}
                )
                
                json_response = response.choices[0].message.content.strip()
                json_response = json_response.replace('```json', '').replace('```', '').strip()
                if json_response.startswith('```'):
                    lines = json_response.split('\n')
                    json_response = '\n'.join(lines[1:-1]) if len(lines) > 2 else json_response
                
                print(f"      📥 Response: {len(json_response):,} chars")
                
                try:
                    batch_extracted = json.loads(json_response)
                    
                    returned_docs = [k for k in batch_extracted.keys() if not k.startswith('_')]
                    missing_docs = set(batch_docs) - set(returned_docs)
                    
                    if missing_docs:
                        print(f"      ⚠️  Missing {len(missing_docs)} doc(s):")
                        for missing in missing_docs:
                            print(f"         - {missing}")
                            doc_type = document_content_map[missing]['type']
                            batch_extracted[missing] = {
                                f"{doc_type}_FILE_DATA": {
                                    "Extraction_Status": {
                                        "value": "Not returned by API (token limit/truncation)",
                                        "confidence": 0.0,
                                        "page": "N/A"
                                    }
                                }
                            }
                    else:
                        print(f"      ✅ All {len(batch_docs)} documents extracted!")
                    
                    all_extracted_data.update(batch_extracted)
                    
                    batch_fields = sum(
                        len(section) 
                        for doc_data in batch_extracted.values() 
                        if not isinstance(doc_data, str)
                        for section in doc_data.values()
                        if isinstance(section, dict)
                    )
                    print(f"      📊 Extracted {batch_fields} fields\n")
                    
                except json.JSONDecodeError as je:
                    print(f"      ❌ JSON parse error: {je}")
                    print(f"      Preview: {json_response[:300]}")
                    
                    for doc_name in batch_docs:
                        doc_type = document_content_map[doc_name]['type']
                        all_extracted_data[doc_name] = {
                            f"{doc_type}_FILE_DATA": {
                                "Extraction_Error": {
                                    "value": f"JSON parse failed: {str(je)[:100]}",
                                    "confidence": 0.0,
                                    "page": "N/A"
                                }
                            }
                        }
            
            except Exception as api_err:
                print(f"      ❌ API error: {type(api_err).__name__}: {str(api_err)[:200]}")
                
                for doc_name in batch_docs:
                    doc_type = document_content_map[doc_name]['type']
                    all_extracted_data[doc_name] = {
                        f"{doc_type}_FILE_DATA": {
                            "API_Error": {
                                "value": f"{type(api_err).__name__}: {str(api_err)[:150]}",
                                "confidence": 0.0,
                                "page": "N/A"
                            }
                        }
                    }
        
        # FINAL SUMMARY
        total_fields = sum(
            len(section)
            for doc_data in all_extracted_data.values()
            if isinstance(doc_data, dict)
            for section in doc_data.values()
            if isinstance(section, dict)
        )
        
        print(f"\n   📊 EXTRACTION COMPLETE:")
        print(f"      Total documents: {len(document_list)}")
        print(f"      Total fields: {total_fields}")
        print(f"      Documents processed:")
        
        for doc_name in document_list:
            if doc_name in all_extracted_data:
                doc_fields = sum(
                    len(section)
                    for section in all_extracted_data[doc_name].values()
                    if isinstance(section, dict)
                )
                status = "✅" if doc_fields >= 15 else "⚠️"
                print(f"         {status} {doc_name}: {doc_fields} fields")
        
        # BUILD FINAL OUTPUT
        final_output = {
            "PROCESSED_FILES": {
                fname: {
                    "file_type": info["file_type"],
                    "is_attachment": info.get("is_attachment", False),
                    **({"has_attachments": True, "attachment_count": info["attachment_count"]} 
                       if info.get("has_attachments") else {}),
                    **({"parent_file": info["parent_file"]} 
                       if info.get("parent_file") else {})
                }
                for fname, info in processed_files_info.items()
            }
        }
        
        # Add extracted data
        for filename in document_list:
            if filename in all_extracted_data:
                final_output[filename] = all_extracted_data[filename]
        
        # Add metadata
        timing_data["ai_extraction_end"] = time.time()
        duration = timing_data["ai_extraction_end"] - timing_data["ai_extraction_start"]
        
        final_output["_EXTRACTION_METADATA"] = {
            "extraction_timestamp": datetime.now().isoformat(),
            "total_documents_processed": len(document_list),
            "extraction_mode": "TOKEN_OPTIMIZED_BATCH",
            "batch_size": batch_size,
            "total_batches": total_batches,
            "api_call_statistics": {
                "document_intelligence_calls": api_call_tracker.get("document_intelligence", 0),
                "azure_openai_calls": api_call_tracker.get("azure_openai", 0),
                "total_api_calls": sum(api_call_tracker.values())
            },
            "processing_time": {
                "extraction_duration_seconds": round(duration, 3)
            },
            "extraction_parameters": {
                "model": deployment_name,
                "temperature": 0.05,
                "max_tokens": 15000,
                "content_limit_per_doc": "10,000 chars",
                "total_content_limit_per_batch": "50,000 chars"
            },
            "optimization_notes": "Token-optimized batch processing with max_tokens=15000 (under 16384 limit)"
        }
        
        print(f"\n✅ Extraction completed in {duration:.3f}s")
        print(f"📦 Total API calls: {api_call_tracker.get('azure_openai', 0)}\n")
        
        return final_output
    
    except Exception as e:
        print(f"❌ Critical Error: {type(e).__name__}: {e}\n")
        import traceback
        traceback.print_exc()
        
        return {
            "_EXTRACTION_ERROR": {
                "error_type": type(e).__name__,
                "error_message": str(e),
                "traceback": traceback.format_exc()
            },
            "_EXTRACTION_METADATA": {
                "extraction_timestamp": datetime.now().isoformat(),
                "extraction_mode": "FAILED",
                "api_call_statistics": {
                    "document_intelligence_calls": api_call_tracker.get("document_intelligence", 0),
                    "azure_openai_calls": api_call_tracker.get("azure_openai", 0),
                    "total_api_calls": sum(api_call_tracker.values())
                }
            }
        }
    
def save_json_output(data, output_dir):
    """Save extracted data as JSON."""
    print_timestamp("Saving JSON output")
    timing_data["json_save_time"] = time.time()
    
    try:
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = output_path / f"extraction_{timestamp}.json"
        
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        
        print(f"✅ JSON saved: {output_file}\n")
        return str(output_file)
    
    except Exception as e:
        print(f"❌ Save error: {e}\n")
        return None

def print_detailed_timing_report():
    """Print comprehensive timing report."""
    if not timing_data["process_start_time"] or not timing_data["process_end_time"]:
        return
    
    total_time = timing_data["process_end_time"] - timing_data["process_start_time"]
    
    print(f"\n{'='*100}")
    print("⏱️  DETAILED TIMING REPORT")
    print(f"{'='*100}")
    
    start_dt = datetime.fromtimestamp(timing_data["process_start_time"])
    end_dt = datetime.fromtimestamp(timing_data["process_end_time"])
    
    print(f"\n📅 START TIME:  {start_dt.strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]}")
    print(f"📅 END TIME:    {end_dt.strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]}")
    print(f"⏱️  TOTAL TIME:  {total_time:.3f} seconds ({total_time/60:.2f} minutes)")
    
    print(f"\n{'='*100}")
    print("STAGE-BY-STAGE BREAKDOWN:")
    print(f"{'='*100}")
    
    stages = [
        ("Folder Scanning", timing_data.get("folder_scan_start"), timing_data.get("folder_scan_end")),
        ("Blob Download", timing_data.get("blob_download_start"), timing_data.get("blob_download_end")),
        ("File Detection", timing_data.get("folder_scan_end"), timing_data.get("file_detection_complete")),
        ("Document Parsing", timing_data.get("parsing_start_time"), timing_data.get("parsing_complete_time")),
        ("Document Intelligence", timing_data.get("doc_intelligence_start"), timing_data.get("doc_intelligence_end")),
        ("AI Field Extraction", timing_data.get("ai_extraction_start"), timing_data.get("ai_extraction_end")),
        ("JSON Saving", timing_data.get("json_save_time"), timing_data.get("process_end_time"))
    ]
    
    for stage_name, start, end in stages:
        if start and end:
            duration = end - start
            percentage = (duration / total_time) * 100
            print(f"{stage_name:<30} {duration:>8.3f}s  ({percentage:>5.1f}%)")
    
    print(f"\n{'='*100}")
    print("PER-FILE PROCESSING TIMES:")
    print(f"{'='*100}")
    
    if timing_data["file_timings"]:
        for filename, times in sorted(timing_data["file_timings"].items()):
            print(f"{filename:<50} {times['duration']:>8.3f}s")
    
    print(f"{'='*100}\n")

def main():
    """Main application logic with interactive blob navigation."""
    print("=" * 100)
    print("Advanced Multi-File Document Extractor with Azure AI + Interactive Blob Storage")
    print("Supports: Local folders, Azure Blob Storage, EML, PDF, DOCX, EXCEL, CSV")
    print("=" * 100)
    print()
    
    timing_data["process_start_time"] = time.time()
    
    env_vars = load_and_validate_env()
    if not env_vars:
        sys.exit(1)
    
    # Display instructions
    print("=" * 100)
    print("INPUT OPTIONS:")
    print("=" * 100)
    print("1. Enter 'blob' for interactive Azure Blob Storage navigation")
    print("2. Local file path: C:\\path\\to\\file.pdf")
    print("3. Local folder path: C:\\path\\to\\folder")
    print("4. Azure Blob Storage URL: https://account.blob.core.windows.net/container/path")
    print("=" * 100)
    print()
    
    # Get input choice
    all_file_paths = []
    temp_blob_files = []
    
    choice = input("Enter 'blob' for interactive navigation or press Enter for manual path entry: ").strip().lower()
    
    if choice == 'blob':
        # Interactive blob navigation
        if BLOB_STORAGE_AVAILABLE and env_vars.get("AZURE_STORAGE_CONNECTION_STRING"):
            blob_selection = interactive_blob_navigation(env_vars["AZURE_STORAGE_CONNECTION_STRING"])
            
            if blob_selection:
                # Convert selected files to blob info format
                blob_info_list = []
                for file in blob_selection['files']:
                    blob_info_list.append({
                        "name": file['name'],
                        "size": file['size'],
                        "container": blob_selection['container'],
                        "full_path": f"blob://{blob_selection['container']}/{file['name']}"
                    })
                
                # Download selected blobs
                if blob_info_list:
                    downloaded_files = download_blobs_to_temp_files(
                        env_vars["AZURE_STORAGE_CONNECTION_STRING"],
                        blob_info_list
                    )
                    
                    for file_info in downloaded_files:
                        all_file_paths.append(file_info["temp_path"])
                        temp_blob_files.append(file_info["temp_path"])
        else:
            print("❌ Blob storage not available or not configured")
            sys.exit(1)
    else:
        # Manual path entry (original behavior)
        print("Enter file/folder/blob paths (one per line, empty line to finish):")
        paths = []
        while True:
            path = input(f"Path {len(paths)+1}: ").strip()
            if not path:
                break
            paths.append(path)
        
        if not paths:
            print("❌ No paths provided")
            sys.exit(1)
        
        # Process paths
        for path in paths:
            path = sanitize_path(path)
            
            if is_blob_storage_url(path):
                if not BLOB_STORAGE_AVAILABLE or not env_vars.get("AZURE_STORAGE_CONNECTION_STRING"):
                    print(f"⚠️  Blob storage not available: {path}")
                    continue
                
                container_name, prefix = parse_blob_path(path)
                
                if not container_name:
                    print(f"⚠️  Could not parse blob path: {path}")
                    continue
                
                blob_files = scan_blob_storage_for_files(
                    env_vars["AZURE_STORAGE_CONNECTION_STRING"],
                    container_name,
                    prefix
                )
                
                if blob_files:
                    downloaded_files = download_blobs_to_temp_files(
                        env_vars["AZURE_STORAGE_CONNECTION_STRING"],
                        blob_files
                    )
                    
                    for file_info in downloaded_files:
                        all_file_paths.append(file_info["temp_path"])
                        temp_blob_files.append(file_info["temp_path"])
            else:
                # Local file or folder
                path_obj = Path(path)
                
                if path_obj.is_dir():
                    folder_files = scan_folder_for_files(path)
                    if folder_files:
                        all_file_paths.extend(folder_files)
                elif path_obj.is_file():
                    all_file_paths.append(path)
                else:
                    print(f"⚠️  Path not found: {path}")
    
    if not all_file_paths:
        print("❌ No valid files found")
        sys.exit(1)
    
    # Display detected files
    display_detected_files(all_file_paths)
    
    # Separate PDF and other files
    pdf_files, other_files = separate_pdf_and_other_files(all_file_paths)
    
    print(f"\n📊 File Separation:")
    print(f"   PDF Files: {len(pdf_files)}")
    print(f"   Other Files: {len(other_files)}")
    
    # Only process non-PDF files with this extractor
    if not other_files:
        print("\n⚠️  No non-PDF files to process with multi_file_extractor")
        print("   PDF files should be processed separately with pdf.py")
        for temp_file in temp_blob_files:
            try:
                os.unlink(temp_file)
            except:
                pass
        sys.exit(0)
    
    # Collect all documents (excluding PDFs)
    all_eml_data, all_documents = collect_all_documents(other_files)
    
    if not all_eml_data and not all_documents:
        print("❌ No documents to process")
        for temp_file in temp_blob_files:
            try:
                os.unlink(temp_file)
            except:
                pass
        sys.exit(1)
    
    # Analyze documents
    extracted_data = analyze_documents_single_call(
        all_eml_data,
        all_documents,
        env_vars["DOC_ENDPOINT"],
        env_vars["DOC_KEY"]
    )
    
    if not extracted_data:
        print("❌ Document analysis failed")
        for temp_file in temp_blob_files:
            try:
                os.unlink(temp_file)
            except:
                pass
        sys.exit(1)
    
    # Extract fields with enhanced detail
    structured_data = extract_fields_with_file_sections(
        extracted_data,
        env_vars["AZURE_OPENAI_ENDPOINT"],
        env_vars["AZURE_OPENAI_KEY"]
    )
    
    if not structured_data:
        print("❌ Field extraction failed")
        for temp_file in temp_blob_files:
            try:
                os.unlink(temp_file)
            except:
                pass
        sys.exit(1)
    
    # Save output
    output_dir = r"C:\Users\isarkar2\OneDrive - DXC Production\Desktop\PDF_Console_App\Output Json"
    output_file = save_json_output(structured_data, output_dir)
    
    # Cleanup temp blob files
    if temp_blob_files:
        print_timestamp("Cleaning up temporary blob files...")
        for temp_file in temp_blob_files:
            try:
                os.unlink(temp_file)
                print(f"   ✅ Deleted: {Path(temp_file).name}")
            except Exception as e:
                print(f"   ⚠️  Could not delete: {Path(temp_file).name}")
        print()
    
    if not output_file:
        sys.exit(1)
    
    timing_data["process_end_time"] = time.time()
    
    # Print detailed timing report
    print_detailed_timing_report()
    
    # Final summary
    print("=" * 100)
    print("✅ PROCESSING COMPLETED SUCCESSFULLY!")
    print("=" * 100)
    print(f"\n📁 Output File: {output_file}")
    
    metadata = structured_data.get("_EXTRACTION_METADATA", {})
    print(f"\n📊 SUMMARY:")
    print(f"   Total Files Processed: {metadata.get('total_documents_processed', 0)}")  
    print(f"   Document Intelligence API Calls: {api_call_tracker['document_intelligence']}")
    print(f"   Azure OpenAI API Calls: {api_call_tracker['azure_openai']}")
    print(f"   Total API Calls: {sum(api_call_tracker.values())}")
    
    if pdf_files:
        print(f"\n⚠️  Note: {len(pdf_files)} PDF file(s) detected but not processed")
        print("   Please process PDF files separately with pdf.py:")
        for idx, pdf_file in enumerate(pdf_files, 1):
            print(f"      Path {idx}: {pdf_file}")
    
    print("=" * 100)

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n⚠️  Process interrupted by user")
        sys.exit(0)
    except Exception as e:
        print(f"\n❌ Unexpected error: {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)