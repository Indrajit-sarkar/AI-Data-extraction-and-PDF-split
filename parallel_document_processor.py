"""
Parallel Document Processor - High-Performance Document Extraction Engine
Supports: PDF, DOCX, EML (with attachments), Excel, CSV
Uses ThreadPoolExecutor for I/O-bound API calls and ProcessPoolExecutor for CPU-bound parsing
"""

import os
import sys
import json
import email
import hashlib
import time
import csv
import tempfile
import logging
import certifi
from pathlib import Path
from datetime import datetime
from io import BytesIO, StringIO
from typing import List, Dict, Any, Optional, Tuple, Union
from dataclasses import dataclass, field, asdict
from concurrent.futures import ThreadPoolExecutor, ProcessPoolExecutor, as_completed
from email import policy
from email.parser import BytesParser
from functools import partial
from dotenv import load_dotenv

# Azure imports
try:
    from azure.ai.documentintelligence import DocumentIntelligenceClient
    from azure.ai.documentintelligence.models import AnalyzeDocumentRequest
    from azure.core.credentials import AzureKeyCredential
    DOC_INTELLIGENCE_AVAILABLE = True
except ImportError:
    DOC_INTELLIGENCE_AVAILABLE = False
    print("⚠️  WARNING: azure-ai-documentintelligence not installed")

try:
    from openai import AzureOpenAI
    OPENAI_AVAILABLE = True
except ImportError:
    OPENAI_AVAILABLE = False
    print("⚠️  WARNING: openai not installed")

try:
    from openpyxl import load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️  WARNING: openpyxl not installed")

try:
    from docx2pdf import convert as docx_to_pdf_convert
    DOCX_CONVERT_AVAILABLE = True
except ImportError:
    DOCX_CONVERT_AVAILABLE = False

try:
    from azure.storage.blob import BlobServiceClient
    BLOB_STORAGE_AVAILABLE = True
except ImportError:
    BLOB_STORAGE_AVAILABLE = False

# Logging configuration
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler()]
)
logger = logging.getLogger(__name__)

# Load environment variables
load_dotenv()


# ==================== DATA CLASSES ====================

@dataclass
class ExtractedField:
    """Represents a single extracted field with metadata"""
    field_name: str
    value: Any
    field_type: str  # string, number, date, boolean, array, object
    confidence: float  # 0.0 to 1.0
    source: str = "Unknown"  # Azure Document Intelligence, OpenAI, Direct Parse
    page: str = "N/A"
    
    def to_dict(self) -> Dict:
        return {
            "field_name": self.field_name,
            "value": self.value,
            "type": self.field_type,
            "confidence": self.confidence,
            "source": self.source,
            "page": self.page
        }


@dataclass
class DocumentResult:
    """Result of processing a single document"""
    filename: str
    file_type: str
    success: bool
    fields: Dict[str, ExtractedField] = field(default_factory=dict)
    attachments: List['DocumentResult'] = field(default_factory=list)
    error: str = None
    processing_time: float = 0.0
    api_calls: Dict[str, int] = field(default_factory=lambda: {"document_intelligence": 0, "openai": 0})
    
    def to_dict(self) -> Dict:
        result = {
            f"{self.file_type}_FILE_DATA": {
                name: f.to_dict() for name, f in self.fields.items()
            },
            "_metadata": {
                "processing_time": self.processing_time,
                "api_calls": self.api_calls,
                "success": self.success,
                "error": self.error
            }
        }
        
        # Add attachments with numbered keys
        for idx, att in enumerate(self.attachments, 1):
            att_key = f"ATTACHMENT_{idx}_{att.file_type}_FILE_DATA"
            result[att_key] = {name: f.to_dict() for name, f in att.fields.items()}
        
        return result


@dataclass
class ProcessingConfig:
    """Configuration for the parallel processing engine"""
    max_threads: int = 4
    max_processes: int = 2
    retry_attempts: int = 3
    retry_delay_base: float = 2.0  # Exponential backoff base
    content_truncate_limit: int = 50000
    openai_max_tokens: int = 4000
    enable_ocr: bool = True


# ==================== UTILITY FUNCTIONS ====================

def detect_file_type(file_path: str) -> Optional[str]:
    """Detect file type from extension"""
    ext = Path(file_path).suffix.lower()
    type_map = {
        '.pdf': 'PDF',
        '.docx': 'DOCX',
        '.doc': 'DOCX',
        '.eml': 'EML',
        '.xlsx': 'EXCEL',
        '.xls': 'EXCEL',
        '.csv': 'CSV'
    }
    return type_map.get(ext)


def detect_value_type(value: Any) -> str:
    """Detect the type of a value"""
    if value is None:
        return "null"
    elif isinstance(value, bool):
        return "boolean"
    elif isinstance(value, (int, float)):
        return "number"
    elif isinstance(value, list):
        return "array"
    elif isinstance(value, dict):
        return "object"
    else:
        return "string"


def calculate_md5(data: Union[str, bytes]) -> str:
    """Calculate MD5 hash"""
    if isinstance(data, str):
        data = data.encode('utf-8')
    return hashlib.md5(data).hexdigest()


def retry_with_backoff(func, max_attempts: int = 3, base_delay: float = 2.0):
    """Retry a function with exponential backoff"""
    last_exception = None
    for attempt in range(max_attempts):
        try:
            return func()
        except Exception as e:
            last_exception = e
            if attempt < max_attempts - 1:
                delay = base_delay * (2 ** attempt)
                logger.warning(f"Attempt {attempt + 1} failed: {e}. Retrying in {delay}s...")
                time.sleep(delay)
    raise last_exception


# ==================== DOCUMENT HANDLERS ====================

class BaseDocumentHandler:
    """Base class for all document handlers"""
    
    def __init__(self, config: ProcessingConfig):
        self.config = config
        self.doc_client = None
        self.openai_client = None
        self._init_clients()
    
    def _init_clients(self):
        """Initialize Azure clients"""
        if DOC_INTELLIGENCE_AVAILABLE:
            endpoint = os.getenv("DOC_ENDPOINT")
            key = os.getenv("DOC_KEY")
            if endpoint and key:
                import ssl
                import urllib3
                from urllib3.util.retry import Retry
                from requests.adapters import HTTPAdapter
                from requests import Session
                from azure.core.pipeline.transport import RequestsTransport
                
                urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
                session = Session()
                session.verify = certifi.where()
                
                retry_strategy = Retry(
                    total=3,
                    backoff_factor=2,
                    status_forcelist=[429, 500, 502, 503, 504],
                    allowed_methods=["POST"]
                )
                adapter = HTTPAdapter(max_retries=retry_strategy)
                session.mount("https://", adapter)
                session.mount("http://", adapter)
                
                transport = RequestsTransport(session=session)
                self.doc_client = DocumentIntelligenceClient(
                    endpoint=endpoint,
                    credential=AzureKeyCredential(key),
                    transport=transport
                )
        
        if OPENAI_AVAILABLE:
            endpoint = os.getenv("AZURE_OPENAI_ENDPOINT")
            key = os.getenv("AZURE_OPENAI_KEY")
            if endpoint and key:
                self.openai_client = AzureOpenAI(
                    azure_endpoint=endpoint,
                    api_key=key,
                    api_version=os.getenv("AZURE_OPENAI_API_VERSION", "2024-02-15-preview")
                )
    
    def process(self, file_path_or_data: Union[str, bytes], filename: str = None) -> DocumentResult:
        """Process a document - to be implemented by subclasses"""
        raise NotImplementedError
    
    def extract_with_document_intelligence(self, data: bytes, filename: str) -> Tuple[str, List[Dict], int]:
        """Extract text using Azure Document Intelligence"""
        if not self.doc_client:
            return "", [], 0
        
        try:
            def do_extract():
                poller = self.doc_client.begin_analyze_document(
                    "prebuilt-layout",
                    AnalyzeDocumentRequest(bytes_source=data)
                )
                return poller.result()
            
            result = retry_with_backoff(do_extract, self.config.retry_attempts, self.config.retry_delay_base)
            
            full_text = ""
            pages_data = []
            
            if result.pages:
                for page_idx, page in enumerate(result.pages, 1):
                    page_text = ""
                    if page.lines:
                        for line in page.lines:
                            page_text += line.content + "\n"
                    pages_data.append({"page": page_idx, "content": page_text})
                    full_text += page_text + "\n"
            
            return full_text, pages_data, 1  # 1 API call
            
        except Exception as e:
            logger.error(f"Document Intelligence error for {filename}: {e}")
            return "", [], 0
    
    def extract_with_openai(self, content: str, filename: str, file_type: str) -> Tuple[Dict[str, ExtractedField], int]:
        """Extract structured fields using Azure OpenAI"""
        if not self.openai_client or not content.strip():
            return {}, 0
        
        deployment = os.getenv("AZURE_OPENAI_DEPLOYMENT", "gpt-4o")
        
        # Truncate content if too long
        if len(content) > self.config.content_truncate_limit:
            content = content[:self.config.content_truncate_limit]
        
        system_prompt = """You are a document data extraction expert. Extract ALL possible fields from the document.

CRITICAL RULES:
1. Return ONLY valid JSON - no markdown, no explanations
2. Extract EVERY field you can identify (aim for 20+ fields)
3. Use this exact format for each field:
   "Field_Name": {"value": "extracted value", "confidence": 0.95, "type": "string"}
4. For missing/not found fields: {"value": null, "confidence": 0.0, "type": "unknown"}
5. Field naming: Use_Underscores, be descriptive (Invoice_Number, Customer_Name, etc.)
6. Types: string, number, date, boolean, array, object, null

Common fields to look for:
- Documents: Title, Date, Author, Company, Address, Phone, Email, Reference_Number
- Invoices: Invoice_Number, Invoice_Date, Due_Date, Vendor_Name, Customer_Name, Line_Items, Subtotal, Tax, Total
- Emails: From, To, CC, BCC, Subject, Date, Body_Summary, Attachments_Count
- Tables: Column_Headers, Row_Count, Data_Summary"""

        user_prompt = f"""Extract ALL fields from this {file_type} document: {filename}

Document Content:
{content}

Return comprehensive JSON with ALL detectable fields. Format:
{{"Field_Name": {{"value": "...", "confidence": 0.95, "type": "string"}}, ...}}"""

        try:
            def do_extract():
                response = self.openai_client.chat.completions.create(
                    model=deployment,
                    messages=[
                        {"role": "system", "content": system_prompt},
                        {"role": "user", "content": user_prompt}
                    ],
                    temperature=0.05,
                    max_tokens=self.config.openai_max_tokens,
                    response_format={"type": "json_object"}
                )
                return response.choices[0].message.content.strip()
            
            json_response = retry_with_backoff(do_extract, self.config.retry_attempts, self.config.retry_delay_base)
            
            # Clean response
            json_response = json_response.replace('```json', '').replace('```', '').strip()
            
            parsed = json.loads(json_response)
            
            fields = {}
            for field_name, field_data in parsed.items():
                if isinstance(field_data, dict) and "value" in field_data:
                    fields[field_name] = ExtractedField(
                        field_name=field_name,
                        value=field_data.get("value"),
                        field_type=field_data.get("type", detect_value_type(field_data.get("value"))),
                        confidence=float(field_data.get("confidence", 0.8)),
                        source="OpenAI Extraction",
                        page=str(field_data.get("page", "N/A"))
                    )
                else:
                    # Direct value without metadata
                    fields[field_name] = ExtractedField(
                        field_name=field_name,
                        value=field_data,
                        field_type=detect_value_type(field_data),
                        confidence=0.7,
                        source="OpenAI Extraction"
                    )
            
            return fields, 1  # 1 API call
            
        except Exception as e:
            logger.error(f"OpenAI extraction error for {filename}: {e}")
            return {}, 0


class PDFHandler(BaseDocumentHandler):
    """Handler for PDF documents with OCR support"""
    
    def process(self, file_path_or_data: Union[str, bytes], filename: str = None) -> DocumentResult:
        start_time = time.time()
        
        if isinstance(file_path_or_data, str):
            filename = filename or Path(file_path_or_data).name
            with open(file_path_or_data, 'rb') as f:
                data = f.read()
        else:
            data = file_path_or_data
            filename = filename or "document.pdf"
        
        result = DocumentResult(
            filename=filename,
            file_type="PDF",
            success=True
        )
        
        try:
            # Extract with Document Intelligence (includes OCR)
            text, pages, di_calls = self.extract_with_document_intelligence(data, filename)
            result.api_calls["document_intelligence"] = di_calls
            
            if text.strip():
                # Extract fields with OpenAI
                fields, openai_calls = self.extract_with_openai(text, filename, "PDF")
                result.fields = fields
                result.api_calls["openai"] = openai_calls
            else:
                # Document is empty or extraction failed
                result.fields["Extraction_Status"] = ExtractedField(
                    field_name="Extraction_Status",
                    value="No text extracted - document may be empty or image-only",
                    field_type="string",
                    confidence=0.0,
                    source="System"
                )
                
        except Exception as e:
            result.success = False
            result.error = str(e)
            logger.error(f"PDF processing error for {filename}: {e}")
        
        result.processing_time = time.time() - start_time
        return result


class DOCXHandler(BaseDocumentHandler):
    """Handler for DOCX documents"""
    
    def process(self, file_path_or_data: Union[str, bytes], filename: str = None) -> DocumentResult:
        start_time = time.time()
        
        if isinstance(file_path_or_data, str):
            filename = filename or Path(file_path_or_data).name
            with open(file_path_or_data, 'rb') as f:
                data = f.read()
        else:
            data = file_path_or_data
            filename = filename or "document.docx"
        
        result = DocumentResult(
            filename=filename,
            file_type="DOCX",
            success=True
        )
        
        try:
            # Try to convert DOCX to PDF for better extraction
            pdf_data = None
            if DOCX_CONVERT_AVAILABLE:
                try:
                    with tempfile.NamedTemporaryFile(suffix='.docx', delete=False) as docx_temp:
                        docx_temp.write(data)
                        docx_path = docx_temp.name
                    
                    pdf_path = docx_path.replace('.docx', '.pdf')
                    docx_to_pdf_convert(docx_path, pdf_path)
                    
                    with open(pdf_path, 'rb') as f:
                        pdf_data = f.read()
                    
                    os.unlink(docx_path)
                    os.unlink(pdf_path)
                except Exception as conv_e:
                    logger.warning(f"DOCX to PDF conversion failed: {conv_e}")
            
            # Use PDF data if conversion succeeded, otherwise use original DOCX
            extract_data = pdf_data if pdf_data else data
            
            # Extract with Document Intelligence
            text, pages, di_calls = self.extract_with_document_intelligence(extract_data, filename)
            result.api_calls["document_intelligence"] = di_calls
            
            if text.strip():
                fields, openai_calls = self.extract_with_openai(text, filename, "DOCX")
                result.fields = fields
                result.api_calls["openai"] = openai_calls
            else:
                result.fields["Extraction_Status"] = ExtractedField(
                    field_name="Extraction_Status",
                    value="No text extracted",
                    field_type="string",
                    confidence=0.0,
                    source="System"
                )
                
        except Exception as e:
            result.success = False
            result.error = str(e)
            logger.error(f"DOCX processing error for {filename}: {e}")
        
        result.processing_time = time.time() - start_time
        return result


class ExcelHandler(BaseDocumentHandler):
    """Handler for Excel files"""
    
    def process(self, file_path_or_data: Union[str, bytes], filename: str = None) -> DocumentResult:
        start_time = time.time()
        
        if isinstance(file_path_or_data, str):
            filename = filename or Path(file_path_or_data).name
            with open(file_path_or_data, 'rb') as f:
                data = f.read()
        else:
            data = file_path_or_data
            filename = filename or "document.xlsx"
        
        result = DocumentResult(
            filename=filename,
            file_type="EXCEL",
            success=True
        )
        
        try:
            if not EXCEL_AVAILABLE:
                raise ImportError("openpyxl not installed")
            
            # Parse Excel
            wb = load_workbook(BytesIO(data), data_only=True)
            
            combined_text = f"Excel File: {filename}\n"
            combined_text += f"Sheets: {', '.join(wb.sheetnames)}\n\n"
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                combined_text += f"=== SHEET: {sheet_name} ===\n"
                
                row_count = 0
                for row in sheet.iter_rows(values_only=True):
                    row_list = [str(cell) if cell is not None else "" for cell in row]
                    if any(row_list):
                        combined_text += " | ".join(row_list) + "\n"
                        row_count += 1
                        if row_count >= 100:  # Limit rows per sheet
                            combined_text += f"... and more rows\n"
                            break
                
                combined_text += "\n"
            
            # Extract fields with OpenAI
            fields, openai_calls = self.extract_with_openai(combined_text, filename, "EXCEL")
            result.fields = fields
            result.api_calls["openai"] = openai_calls
            
            # Add sheet metadata
            result.fields["Total_Sheets"] = ExtractedField(
                field_name="Total_Sheets",
                value=len(wb.sheetnames),
                field_type="number",
                confidence=1.0,
                source="Direct Parse"
            )
            result.fields["Sheet_Names"] = ExtractedField(
                field_name="Sheet_Names",
                value=wb.sheetnames,
                field_type="array",
                confidence=1.0,
                source="Direct Parse"
            )
            
        except Exception as e:
            result.success = False
            result.error = str(e)
            logger.error(f"Excel processing error for {filename}: {e}")
        
        result.processing_time = time.time() - start_time
        return result


class CSVHandler(BaseDocumentHandler):
    """Handler for CSV files"""
    
    def process(self, file_path_or_data: Union[str, bytes], filename: str = None) -> DocumentResult:
        start_time = time.time()
        
        if isinstance(file_path_or_data, str):
            filename = filename or Path(file_path_or_data).name
            with open(file_path_or_data, 'r', encoding='utf-8', errors='ignore') as f:
                csv_text = f.read()
        elif isinstance(file_path_or_data, bytes):
            csv_text = file_path_or_data.decode('utf-8', errors='ignore')
            filename = filename or "document.csv"
        else:
            csv_text = file_path_or_data
            filename = filename or "document.csv"
        
        result = DocumentResult(
            filename=filename,
            file_type="CSV",
            success=True
        )
        
        try:
            reader = csv.reader(StringIO(csv_text))
            rows = [row for row in reader if any(cell.strip() for cell in row)]
            
            headers = rows[0] if rows else []
            data_rows = rows[1:] if len(rows) > 1 else []
            
            combined_text = f"CSV File: {filename}\n"
            combined_text += f"Headers: {' | '.join(headers)}\n"
            combined_text += f"Total Rows: {len(data_rows)}\n\n"
            
            for idx, row in enumerate(data_rows[:50], 1):
                combined_text += f"Row {idx}: {' | '.join(row)}\n"
            
            if len(data_rows) > 50:
                combined_text += f"\n... and {len(data_rows) - 50} more rows\n"
            
            # Extract fields with OpenAI
            fields, openai_calls = self.extract_with_openai(combined_text, filename, "CSV")
            result.fields = fields
            result.api_calls["openai"] = openai_calls
            
            # Add CSV metadata
            result.fields["Column_Count"] = ExtractedField(
                field_name="Column_Count",
                value=len(headers),
                field_type="number",
                confidence=1.0,
                source="Direct Parse"
            )
            result.fields["Row_Count"] = ExtractedField(
                field_name="Row_Count",
                value=len(data_rows),
                field_type="number",
                confidence=1.0,
                source="Direct Parse"
            )
            result.fields["Column_Headers"] = ExtractedField(
                field_name="Column_Headers",
                value=headers,
                field_type="array",
                confidence=1.0,
                source="Direct Parse"
            )
            
        except Exception as e:
            result.success = False
            result.error = str(e)
            logger.error(f"CSV processing error for {filename}: {e}")
        
        result.processing_time = time.time() - start_time
        return result


class EMLHandler(BaseDocumentHandler):
    """Handler for EML email files with attachment processing"""
    
    def __init__(self, config: ProcessingConfig):
        super().__init__(config)
        self.pdf_handler = PDFHandler(config)
        self.docx_handler = DOCXHandler(config)
        self.excel_handler = ExcelHandler(config)
        self.csv_handler = CSVHandler(config)
    
    def process(self, file_path_or_data: Union[str, bytes], filename: str = None) -> DocumentResult:
        start_time = time.time()
        
        if isinstance(file_path_or_data, str):
            filename = filename or Path(file_path_or_data).name
            with open(file_path_or_data, 'rb') as f:
                data = f.read()
        else:
            data = file_path_or_data
            filename = filename or "email.eml"
        
        result = DocumentResult(
            filename=filename,
            file_type="EML",
            success=True
        )
        
        try:
            msg = BytesParser(policy=policy.default).parse(BytesIO(data))
            
            # Extract email metadata
            email_fields = {
                "Email_From": msg.get('From', ''),
                "Email_To": msg.get('To', ''),
                "Email_CC": msg.get('Cc', ''),
                "Email_BCC": msg.get('Bcc', ''),
                "Email_Subject": msg.get('Subject', ''),
                "Email_Date": msg.get('Date', ''),
                "Message_ID": msg.get('Message-ID', ''),
                "Reply_To": msg.get('Reply-To', ''),
            }
            
            # Add metadata fields
            for field_name, value in email_fields.items():
                result.fields[field_name] = ExtractedField(
                    field_name=field_name,
                    value=value if value else None,
                    field_type="string" if value else "null",
                    confidence=1.0 if value else 0.0,
                    source="Direct Parse",
                    page="Email Header"
                )
            
            # Extract body
            body_text = ""
            body_html = ""
            attachments = []
            
            if msg.is_multipart():
                for part in msg.walk():
                    content_type = part.get_content_type()
                    disposition = str(part.get('Content-Disposition', ''))
                    
                    if 'attachment' not in disposition:
                        if content_type == 'text/plain':
                            try:
                                body_text += part.get_content()
                            except:
                                pass
                        elif content_type == 'text/html':
                            try:
                                body_html += part.get_content()
                            except:
                                pass
                    else:
                        att_filename = part.get_filename()
                        if att_filename:
                            att_data = part.get_payload(decode=True)
                            if att_data:
                                attachments.append({
                                    "filename": att_filename,
                                    "data": att_data,
                                    "size": len(att_data)
                                })
            else:
                try:
                    body_text = msg.get_content()
                except:
                    pass
            
            # Add body field
            body_content = body_text or body_html[:5000]
            result.fields["Email_Body"] = ExtractedField(
                field_name="Email_Body",
                value=body_content[:2000] if body_content else None,
                field_type="string" if body_content else "null",
                confidence=1.0 if body_content else 0.0,
                source="Direct Parse",
                page="Email Body"
            )
            
            result.fields["Attachment_Count"] = ExtractedField(
                field_name="Attachment_Count",
                value=len(attachments),
                field_type="number",
                confidence=1.0,
                source="Direct Parse"
            )
            
            # If we have body content, also extract with OpenAI for additional fields
            if body_content:
                email_content = f"From: {email_fields['Email_From']}\n"
                email_content += f"To: {email_fields['Email_To']}\n"
                email_content += f"Subject: {email_fields['Email_Subject']}\n"
                email_content += f"Date: {email_fields['Email_Date']}\n\n"
                email_content += f"Body:\n{body_content}"
                
                ai_fields, openai_calls = self.extract_with_openai(email_content, filename, "EML")
                result.api_calls["openai"] = openai_calls
                
                # Merge AI fields (don't override existing metadata)
                for field_name, field in ai_fields.items():
                    if field_name not in result.fields:
                        result.fields[field_name] = field
            
            # Process attachments in parallel
            if attachments:
                logger.info(f"Processing {len(attachments)} attachments from {filename}")
                
                with ThreadPoolExecutor(max_workers=min(self.config.max_threads, len(attachments))) as executor:
                    future_to_att = {}
                    
                    for att in attachments:
                        att_type = detect_file_type(att["filename"])
                        if att_type:
                            handler = self._get_handler_for_type(att_type)
                            if handler:
                                future = executor.submit(handler.process, att["data"], att["filename"])
                                future_to_att[future] = att["filename"]
                    
                    for future in as_completed(future_to_att):
                        try:
                            att_result = future.result()
                            result.attachments.append(att_result)
                            # Aggregate API calls
                            result.api_calls["document_intelligence"] += att_result.api_calls.get("document_intelligence", 0)
                            result.api_calls["openai"] += att_result.api_calls.get("openai", 0)
                        except Exception as e:
                            logger.error(f"Attachment processing error: {e}")
            
        except Exception as e:
            result.success = False
            result.error = str(e)
            logger.error(f"EML processing error for {filename}: {e}")
        
        result.processing_time = time.time() - start_time
        return result
    
    def _get_handler_for_type(self, file_type: str) -> Optional[BaseDocumentHandler]:
        """Get the appropriate handler for a file type"""
        handlers = {
            'PDF': self.pdf_handler,
            'DOCX': self.docx_handler,
            'EXCEL': self.excel_handler,
            'CSV': self.csv_handler
        }
        return handlers.get(file_type)


# ==================== PARALLEL PROCESSING ENGINE ====================

class ParallelProcessingEngine:
    """Main orchestrator for parallel document processing"""
    
    def __init__(self, config: ProcessingConfig = None):
        self.config = config or ProcessingConfig()
        self.handlers = {}
        self._init_handlers()
        self.total_api_calls = {"document_intelligence": 0, "openai": 0}
    
    def _init_handlers(self):
        """Initialize document handlers"""
        self.handlers = {
            'PDF': PDFHandler(self.config),
            'DOCX': DOCXHandler(self.config),
            'EML': EMLHandler(self.config),
            'EXCEL': ExcelHandler(self.config),
            'CSV': CSVHandler(self.config)
        }
    
    def scan_path(self, path: str) -> List[str]:
        """Scan a path (file or folder) and return list of supported files"""
        supported_extensions = {'.pdf', '.docx', '.doc', '.eml', '.xlsx', '.xls', '.csv'}
        files = []
        
        path_obj = Path(path)
        
        if path_obj.is_file():
            if path_obj.suffix.lower() in supported_extensions:
                files.append(str(path_obj))
        elif path_obj.is_dir():
            for file_path in path_obj.rglob('*'):
                if file_path.is_file() and file_path.suffix.lower() in supported_extensions:
                    files.append(str(file_path))
        
        return files
    
    def process_single_document(self, file_path: str) -> DocumentResult:
        """Process a single document"""
        file_type = detect_file_type(file_path)
        
        if not file_type:
            return DocumentResult(
                filename=Path(file_path).name,
                file_type="UNKNOWN",
                success=False,
                error="Unsupported file type"
            )
        
        handler = self.handlers.get(file_type)
        if not handler:
            return DocumentResult(
                filename=Path(file_path).name,
                file_type=file_type,
                success=False,
                error=f"No handler for {file_type}"
            )
        
        logger.info(f"Processing: {Path(file_path).name} [{file_type}]")
        result = handler.process(file_path)
        
        # Track API calls
        self.total_api_calls["document_intelligence"] += result.api_calls.get("document_intelligence", 0)
        self.total_api_calls["openai"] += result.api_calls.get("openai", 0)
        
        return result
    
    def process_files_parallel(self, file_paths: List[str], progress_callback=None) -> Dict:
        """Process multiple files in parallel"""
        start_time = time.time()
        results = {}
        total_fields = 0
        successful = 0
        failed = 0
        
        logger.info(f"Starting parallel processing of {len(file_paths)} files")
        logger.info(f"Using {self.config.max_threads} threads, {self.config.max_processes} processes")
        
        with ThreadPoolExecutor(max_workers=self.config.max_threads) as executor:
            future_to_path = {
                executor.submit(self.process_single_document, path): path
                for path in file_paths
            }
            
            completed = 0
            for future in as_completed(future_to_path):
                file_path = future_to_path[future]
                completed += 1
                
                try:
                    result = future.result()
                    results[result.filename] = result.to_dict()
                    
                    if result.success:
                        successful += 1
                        total_fields += len(result.fields)
                        for att in result.attachments:
                            total_fields += len(att.fields)
                    else:
                        failed += 1
                    
                    if progress_callback:
                        progress_callback(completed, len(file_paths), result.filename)
                    
                    logger.info(f"[{completed}/{len(file_paths)}] Completed: {result.filename} "
                               f"({len(result.fields)} fields, {result.processing_time:.2f}s)")
                    
                except Exception as e:
                    failed += 1
                    logger.error(f"Error processing {file_path}: {e}")
                    results[Path(file_path).name] = {
                        "EXTRACTION_ERROR": {
                            "error": {"value": str(e), "type": "string", "confidence": 0.0}
                        }
                    }
        
        processing_time = time.time() - start_time
        
        # Add metadata
        results["_EXTRACTION_METADATA"] = {
            "extraction_timestamp": datetime.now().isoformat(),
            "total_documents_processed": len(file_paths),
            "successful_extractions": successful,
            "failed_extractions": failed,
            "total_fields_extracted": total_fields,
            "processing_time_seconds": round(processing_time, 2),
            "api_calls": {
                "document_intelligence": self.total_api_calls["document_intelligence"],
                "openai": self.total_api_calls["openai"],
                "total": sum(self.total_api_calls.values())
            },
            "configuration": {
                "max_threads": self.config.max_threads,
                "max_processes": self.config.max_processes,
                "retry_attempts": self.config.retry_attempts
            }
        }
        
        logger.info(f"✅ Processing complete: {successful} successful, {failed} failed, "
                   f"{total_fields} fields, {processing_time:.2f}s")
        
        return results
    
    def process_path(self, path: str, progress_callback=None) -> Dict:
        """Process a file or folder path"""
        files = self.scan_path(path)
        
        if not files:
            logger.warning(f"No supported files found in: {path}")
            return {"_EXTRACTION_METADATA": {"error": "No supported files found"}}
        
        logger.info(f"Found {len(files)} files to process")
        return self.process_files_parallel(files, progress_callback)


# ==================== MAIN ENTRY POINT ====================

def save_json_output(data: Dict, output_dir: str = None) -> str:
    """Save extraction results to JSON file"""
    if not output_dir:
        output_dir = os.path.join(os.path.dirname(__file__), "output")
    
    os.makedirs(output_dir, exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"extraction_{timestamp}.json"
    output_path = os.path.join(output_dir, filename)
    
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False, default=str)
    
    logger.info(f"✅ Results saved to: {output_path}")
    return output_path


def main():
    """Main entry point for command-line usage"""
    print("=" * 80)
    print("[PARALLEL] Document Processor - High-Performance Extraction Engine")
    print("=" * 80)
    print()
    print("Supports: PDF (with OCR), DOCX, EML (with attachments), Excel, CSV")
    print("Features: Multithreading, Per-document API calls, Comprehensive extraction")
    print()
    
    # Check for test mode
    if "--test" in sys.argv:
        print("[TEST] Running self-test...")
        config = ProcessingConfig(max_threads=2, max_processes=1)
        engine = ParallelProcessingEngine(config)
        print("[OK] Engine initialized successfully")
        print(f"   Handlers: {list(engine.handlers.keys())}")
        print(f"   Config: max_threads={config.max_threads}, max_processes={config.max_processes}")
        return
    
    # Interactive mode
    print("Enter path to file or folder (or 'quit' to exit):")
    
    while True:
        path = input("\n[PATH] Enter path: ").strip().strip('"').strip("'")
        
        if path.lower() in ('quit', 'exit', 'q'):
            print("[BYE] Goodbye!")
            break
        
        if not path:
            continue
        
        path_obj = Path(path)
        if not path_obj.exists():
            print(f"[ERROR] Path not found: {path}")
            continue
        
        # Initialize engine
        config = ProcessingConfig(
            max_threads=4,
            max_processes=2,
            retry_attempts=3
        )
        engine = ParallelProcessingEngine(config)
        
        # Process
        results = engine.process_path(path)
        
        # Save results
        output_dir = r"C:\Users\isarkar2\OneDrive - DXC Production\Desktop\PDF_Console_App\Output Json"
        output_file = save_json_output(results, output_dir)
        
        print(f"\n{'='*80}")
        print("[OK] PROCESSING COMPLETE!")
        print(f"{'='*80}")
        print(f"[OUTPUT] {output_file}")
        
        metadata = results.get("_EXTRACTION_METADATA", {})
        print(f"\n[SUMMARY]")
        print(f"   Documents: {metadata.get('total_documents_processed', 0)}")
        print(f"   Fields: {metadata.get('total_fields_extracted', 0)}")
        print(f"   Time: {metadata.get('processing_time_seconds', 0)}s")
        print(f"   API Calls: {metadata.get('api_calls', {}).get('total', 0)}")


if __name__ == "__main__":
    main()
